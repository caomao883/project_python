#coding:utf-8

from openpyxl import load_workbook,Workbook
wb = load_workbook("data/test.xlsx")

print(wb.sheetnames)


sheet = wb.get_sheet_by_name("Sheet2")

print(sheet["C1"].value) #获取单元的值

print(sheet.max_row)    # 10     <-最大行数
print(sheet.max_column)    # 5     <-最大列数



#遍历第一行
for i in sheet["1"]:
    print(i.value)

#遍历C列
for i in sheet["C"]:
  print(i.value)


new_sheet = wb.create_sheet()
new_sheet.title = "Sheet3"
new_sheet["C3"]=45
print(wb.sheetnames)
#保存修改
wb.save("data/test.xlsx")

