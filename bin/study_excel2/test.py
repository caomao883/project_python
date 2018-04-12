# coding:utf-8

from openpyxl import load_workbook, Workbook

wb = load_workbook("data/tt.xlsx")

print(wb.sheetnames)


def func(value):
    value = value.strip("\n\r")
    result = ""
    i = 0
    n = len(value)
    temp = ""
    while i < n:
        if value[i] >= "0" and value[i] <= "9":
            temp += value[i]
        elif value[i] == 'm':
            if value[i + 1] == "l":
                temp += "ml"
                result +=temp
                break
        else:
            break
        i+=1
    temp = ""
    if value.endswith("mg"):
        temp = "mg"
        i = n-3
        while i >=0:
            if value[i] >= "0" and value[i] <= "9":
                temp = value[i] + temp
            else:
                break
            i-=1
    if len(temp) >0:
        result += temp
    return result


sheet = wb.get_sheet_by_name("Sheet1")
# 遍历C列
for rowOfCellObjects in sheet['C2':'C233']:
    print("")
    # for cellObj in rowOfCellObjects:
    #     print(cellObj.coordinate, cellObj.value)
    #     sheet["F"+cellObj.coordinate[1:]] = func(cellObj.value)
wb.save("data/tt.xlsx")
#
# print(sheet["C1"].value) #获取单元的值
#
# print(sheet.max_row)    # 10     <-最大行数
# print(sheet.max_column)    # 5     <-最大列数
#
#
#
# #遍历第一行
# for i in sheet["1"]:
#     print(i.value)
#
# #遍历C列
# for i in sheet["C"]:
#   print(i.value)
#
#
# new_sheet = wb.create_sheet()
# new_sheet.title = "Sheet3"
# new_sheet["C3"]=45
# print(wb.sheetnames)
# #保存修改
wb.save("data/test.xlsx")
#
